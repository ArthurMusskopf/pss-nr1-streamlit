from __future__ import annotations

import textwrap
import base64
import csv
import json
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from html import escape
from io import BytesIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import streamlit as st
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

APP_DIR = Path(__file__).parent
DATA_DIR = APP_DIR / "data"
ASSETS_DIR = APP_DIR / "assets"
LOGO_SYMBOL_WHITE = ASSETS_DIR / "logo_symbol_white.png"
LOGO_WORDMARK_WHITE = ASSETS_DIR / "logo_wordmark_white.png"
TZ_BR = ZoneInfo("America/Sao_Paulo")

BRAND = {
    "graphite": "#595C65",
    "sand": "#D2C3A0",
    "mist": "#EDEDEC",
    "warm_gray": "#A8A5A2",
    "green": "#2B3928",
    "black": "#000000",
    "white": "#FFFFFF",
    "terracotta": "#B47250",
    "dark_gray": "#4D4D4D",
    "success": "#2B5A35",
    "danger": "#8B3D2E",
}

WORKSHOP = {
    "titulo": "Workshop NR-1: Gestão de Riscos Psicossociais",
    "subtitulo": "Checklist de Adequação à NR-1",
    "carga_horaria": "4 horas",
    "data": "29/05/2026",
    "local": "São Paulo, SP",
    "assinatura_1_nome": "Carolina Perroni Sanvicene",
    "assinatura_1_cargo": "Sócia | PS&S Advogados",
    "assinatura_2_nome": "Simone Menda",
    "assinatura_2_cargo": "Sócia | Entre Telas e Agora",
}

SCORES = {
    "Sim - Conforme": 5.0,
    "Em andamento": 2.5,
    "Não - Inconformidade": 0.0,
    "N/A": None,
}
STATUS_OPTIONS = list(SCORES.keys())


@dataclass
class Question:
    numero: int
    area: str
    requisito: str
    pergunta: str


@dataclass
class Diagnosis:
    pontos: float
    maximo: float
    percentual: float
    status: str
    status_label: str
    descricao: str
    pacote: str
    cor: str


def load_questions() -> list[Question]:
    path = DATA_DIR / "perguntas.csv"
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [
            Question(
                numero=int(row["numero"]),
                area=row["area"],
                requisito=row["requisito"],
                pergunta=row["pergunta"],
            )
            for row in reader
        ]


def asset_b64(path: Path) -> str:
    if not path.exists():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


def css() -> None:
    """CSS completo e isolado do app.

    Importante: esta função NÃO lê assets/styles.css. O arquivo de CSS do repositório
    tinha overrides e hotfixes conflitantes para o header. Centralizar o estilo aqui
    deixa a renderização previsível no Streamlit Cloud.
    """
    style = """
@import url('https://fonts.googleapis.com/css2?family=Lato:wght@400;700;900&display=swap');

:root {
    --graphite: #595C65;
    --sand: #D2C3A0;
    --mist: #EDEDEC;
    --warm-gray: #A8A5A2;
    --green: #2B3928;
    --success: #2B5A35;
    --danger: #8B3D2E;
    --terracotta: #B47250;
    --black: #000000;
    --white: #FFFFFF;
    --line: rgba(0, 0, 0, .12);
    --shadow: 0 14px 34px rgba(0, 0, 0, .10);
}

* {
    box-sizing: border-box;
    font-family: Lato, Arial, sans-serif;
}

html,
body,
.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background: var(--white) !important;
    color: var(--black) !important;
}

[data-testid="stHeader"],
[data-testid="stDecoration"],
[data-testid="stToolbar"],
#MainMenu,
footer {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
}

.block-container,
section.main > div {
    max-width: 1180px !important;
    margin: 0 auto !important;
    padding: 0 28px 2.5rem 28px !important;
}

p,
label,
span,
li,
div,
h1,
h2,
h3,
h4,
h5,
h6 {
    color: var(--black);
}

/* Header */
.pss-hero {
    width: 100vw;
    margin-left: calc(50% - 50vw);
    margin-right: calc(50% - 50vw);
    margin-top: -18px;
    background: var(--graphite);
    box-shadow: 0 10px 28px rgba(0, 0, 0, .16);
}

.pss-hero-inner {
    width: min(1180px, 100%);
    min-height: 220px;
    margin: 0 auto;
    padding: 44px 40px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 52px;
}

.pss-brand-lockup {
    flex: 1 1 auto;
    min-width: 0;
    display: flex;
    align-items: center;
    gap: 28px;
}

.pss-logo-symbol {
    width: 62px;
    min-width: 62px;
    height: auto;
    display: block;
}

.pss-logo-wordmark {
    width: min(520px, 100%);
    max-height: 132px;
    object-fit: contain;
    display: block;
}

.pss-wordmark-fallback {
    color: var(--mist) !important;
    font-size: clamp(2.1rem, 5vw, 4.2rem);
    line-height: .95;
    font-weight: 900;
    letter-spacing: -.06em;
    text-transform: uppercase;
}

.pss-hero-copy {
    flex: 0 0 410px;
    text-align: right;
    min-width: 320px;
}

.pss-hero-title {
    color: var(--sand) !important;
    font-size: clamp(1.55rem, 2.4vw, 2.25rem);
    line-height: 1.08;
    font-weight: 900;
    letter-spacing: -.025em;
    margin: 0 0 12px 0;
}

.pss-hero-copy p {
    color: var(--mist) !important;
    font-size: 1rem;
    line-height: 1.45;
    font-weight: 700;
    margin: 0;
}

/* Main layout */
.pss-content {
    max-width: 1120px;
    margin: 0 auto;
    padding: 42px 0 0 0;
}

.pss-card,
[data-testid="stForm"],
[data-testid="stVerticalBlockBorderWrapper"] {
    background: var(--white) !important;
    background-color: var(--white) !important;
    border: 1px solid var(--line) !important;
    border-radius: 14px !important;
    box-shadow: var(--shadow) !important;
}

.pss-card {
    padding: 22px 24px;
    margin: 0 0 22px 0;
}

[data-testid="stForm"] {
    padding: 22px 24px !important;
    margin-bottom: 22px !important;
}

[data-testid="stVerticalBlockBorderWrapper"] {
    padding: 18px 20px !important;
    margin: 0 0 18px 0 !important;
}

[data-testid="stVerticalBlockBorderWrapper"] > * {
    background: transparent !important;
}

.pss-card-title {
    margin: 0 0 16px 0;
    padding-bottom: 10px;
    border-bottom: 1px solid rgba(0, 0, 0, .14);
    letter-spacing: .16em;
    text-transform: uppercase;
    font-size: .76rem;
    font-weight: 900;
    color: var(--black) !important;
}

.pss-intro {
    border-left: 6px solid var(--sand);
    font-size: 1rem;
    line-height: 1.72;
}

.pss-card p,
.pss-card strong,
.pss-card li {
    color: var(--black) !important;
}

.pss-card strong {
    font-weight: 900;
}

.small-note {
    color: rgba(0, 0, 0, .62) !important;
    font-size: .92rem;
}

/* Identification */
.pss-company-card {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 12px 20px;
}

.pss-company-item .label {
    display: block;
    color: rgba(0, 0, 0, .58) !important;
    font-size: .74rem;
    text-transform: uppercase;
    letter-spacing: .12em;
    font-weight: 900;
}

.pss-company-item .value {
    display: block;
    color: var(--black) !important;
    font-size: .98rem;
    font-weight: 700;
}

/* Streamlit controls */
.stRadio > label,
.stTextInput > label,
.stTextArea > label,
.stCheckbox > label {
    color: var(--black) !important;
    font-weight: 800 !important;
}

div[role="radiogroup"] label p,
.stCheckbox label p {
    color: var(--black) !important;
}

.stTextInput input,
.stTextArea textarea {
    background: var(--white) !important;
    color: var(--black) !important;
    border-radius: 8px !important;
    border: 1px solid rgba(0, 0, 0, .24) !important;
}

.stTextInput input:focus,
.stTextArea textarea:focus {
    border-color: var(--success) !important;
    box-shadow: 0 0 0 1px var(--success) !important;
}

div.stButton > button,
div.stDownloadButton > button {
    background: var(--success) !important;
    border-color: var(--success) !important;
    color: var(--mist) !important;
    font-weight: 900 !important;
    min-width: 240px;
}

/* Questions */
.pss-section-header {
    background: var(--graphite) !important;
    border-radius: 12px 12px 0 0;
    padding: 14px 20px;
    margin-top: 24px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 16px;
}

.pss-section-header .name {
    color: var(--mist) !important;
    text-transform: uppercase;
    letter-spacing: .17em;
    font-size: .76rem;
    font-weight: 900;
}

.pss-section-header .count {
    color: var(--sand) !important;
    font-size: .9rem;
    white-space: nowrap;
    font-weight: 900;
}

.pss-question-meta {
    display: flex;
    gap: 10px;
    align-items: center;
    margin-bottom: 10px;
}

.pss-question-num {
    background: var(--graphite) !important;
    color: var(--mist) !important;
    width: 30px;
    height: 30px;
    border-radius: 50%;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-weight: 900;
}

.pss-question-req {
    color: var(--sand) !important;
    background: var(--graphite) !important;
    padding: 6px 10px;
    border-radius: 999px;
    text-transform: uppercase;
    letter-spacing: .09em;
    font-size: .68rem;
    font-weight: 800;
}

.pss-question-text {
    color: var(--black) !important;
    font-size: 1.02rem;
    line-height: 1.5;
    margin: 0 0 10px 0;
}

/* Progress and result */
.pss-progress,
.pss-result {
    background: var(--white) !important;
    border-radius: 12px;
    border: 1px solid rgba(0, 0, 0, .10);
}

.pss-progress {
    padding: 13px 16px;
    margin: 18px 0;
    display: flex;
    justify-content: space-between;
    font-weight: 800;
}

.pss-result {
    border: 2px solid var(--green);
    padding: 28px 24px 30px;
    margin: 24px auto 18px;
    text-align: center;
    box-shadow: var(--shadow);
}

.pss-result-kicker {
    color: var(--green) !important;
    font-size: .78rem;
    font-weight: 900;
    letter-spacing: .18em;
    text-transform: uppercase;
    margin-bottom: 12px;
}

.pss-result-points {
    color: var(--green) !important;
    font-family: Georgia, 'Times New Roman', serif !important;
    font-size: clamp(3.2rem, 7vw, 5rem);
    font-weight: 900;
    line-height: 1;
}

.pss-result-max {
    color: rgba(0, 0, 0, .56) !important;
    font-weight: 800;
}

.pss-result-bar {
    width: min(540px, 80%);
    height: 12px;
    border-radius: 999px;
    background: rgba(0, 0, 0, .07);
    margin: 22px auto 20px;
    overflow: hidden;
}

.pss-result-bar-fill {
    height: 100%;
    width: var(--value);
    background: var(--status-color);
}

.pss-status-pill {
    display: inline-flex;
    justify-content: center;
    min-width: 108px;
    padding: 9px 24px;
    border-radius: 999px;
    color: var(--status-color) !important;
    border: 1px solid currentColor;
    font-weight: 900;
    text-transform: uppercase;
}

.pss-result-copy {
    max-width: 650px;
    margin: 18px auto 0;
    line-height: 1.58;
}

.pss-cert-id {
    color: rgba(0, 0, 0, .56) !important;
    margin-top: 14px;
    font-size: .9rem;
    font-weight: 800;
}

/* Footer */
.pss-footer-bleed {
    width: 100vw;
    margin-left: calc(50% - 50vw);
    margin-right: calc(50% - 50vw);
    margin-top: 42px;
    background: var(--graphite);
}

.pss-footer-inner {
    max-width: 1180px;
    margin: 0 auto;
    padding: 30px 24px;
    text-align: center;
}

.pss-footer-inner * {
    color: var(--mist) !important;
}

.pss-footer-inner .brand {
    font-weight: 900;
    letter-spacing: .08em;
    text-transform: uppercase;
}

.pss-footer-inner .muted {
    color: var(--sand) !important;
    margin-top: 6px;
    font-size: .9rem;
}

/* Responsive */
@media (max-width: 980px) {
    .block-container,
    section.main > div {
        padding: 0 18px 2rem 18px !important;
    }

    .pss-hero-inner {
        min-height: auto;
        padding: 34px 24px;
        flex-direction: column;
        align-items: flex-start;
        gap: 26px;
    }

    .pss-brand-lockup {
        width: 100%;
        gap: 20px;
    }

    .pss-logo-symbol {
        width: 50px;
        min-width: 50px;
    }

    .pss-logo-wordmark {
        width: min(430px, calc(100vw - 120px));
        max-height: 106px;
    }

    .pss-hero-copy {
        flex: 0 1 auto;
        width: 100%;
        min-width: 0;
        text-align: left;
    }

    .pss-company-card {
        grid-template-columns: 1fr;
    }
}
"""
    st.markdown(f"<style>{style}</style>", unsafe_allow_html=True)


def get_bq_client() -> bigquery.Client | None:
    try:
        bq_conf = st.secrets["bigquery"]
        svc = dict(st.secrets["gcp_service_account"])
    except Exception:
        return None

    try:
        credentials = service_account.Credentials.from_service_account_info(svc)
        return bigquery.Client(credentials=credentials, project=bq_conf["project_id"])
    except Exception:
        return None


def table_id(table_name: str) -> str:
    bq_conf = st.secrets["bigquery"]
    return f"{bq_conf['project_id']}.{bq_conf['dataset_id']}.{table_name}"


def save_local_json(folder: str, filename: str, payload: dict[str, Any]) -> None:
    out = APP_DIR / folder
    out.mkdir(exist_ok=True)
    with (out / filename).open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)


def save_start_record(participant: dict[str, str], start_id: str, created_at: datetime, created_at_br: str) -> tuple[bool, str]:
    row = {
        "start_id": start_id,
        "created_at": created_at.isoformat(),
        "created_at_br": created_at_br,
        "nome": participant.get("nome", ""),
        "email": participant.get("email", ""),
        "empresa": participant.get("empresa", ""),
        "cnpj": participant.get("cnpj", ""),
        "cargo": participant.get("cargo", ""),
        "telefone": participant.get("telefone", ""),
        "aceite_dados": True,
        "fonte": "app_streamlit",
    }
    client = get_bq_client()
    if client is None:
        save_local_json(".local_starts", f"{start_id}.json", row)
        return False, "local_backup"
    errors = client.insert_rows_json(table_id("inicios"), [row])
    if errors:
        save_local_json(".local_starts", f"{start_id}.json", row)
        return False, "remote_error"
    return True, "ok"


def save_to_bigquery(submission: dict[str, Any], answers: list[dict[str, Any]]) -> tuple[bool, str]:
    client = get_bq_client()
    if client is None:
        save_local_json(".local_submissions", f"{submission['submission_id']}.json", {"submission": submission, "answers": answers})
        return False, "local_backup"

    errors_sub = client.insert_rows_json(table_id("submissoes"), [submission])
    errors_ans = client.insert_rows_json(table_id("respostas"), answers)

    if errors_sub or errors_ans:
        save_local_json(".local_submissions", f"{submission['submission_id']}.json", {"submission": submission, "answers": answers})
        return False, "remote_error"
    return True, "ok"


def calculate(answers: list[dict[str, Any]]) -> Diagnosis:
    pontos = 0.0
    maximo = 0.0
    for a in answers:
        score = SCORES[a["resposta"]]
        if score is None:
            continue
        pontos += score
        maximo += 5.0

    percentual = round((pontos / maximo * 100), 0) if maximo else 0.0
    if percentual >= 71:
        return Diagnosis(
            pontos=pontos,
            maximo=maximo,
            percentual=percentual,
            status="ADEQUADO",
            status_label="Adequado",
            descricao="Bom nível de conformidade. Mantenha evidências, atualizações periódicas e monitoramento dos indicadores.",
            pacote="Manutenção e Monitoramento NR-1",
            cor=BRAND["success"],
        )
    if percentual >= 41:
        return Diagnosis(
            pontos=pontos,
            maximo=maximo,
            percentual=percentual,
            status="EM ADEQUAÇÃO",
            status_label="Em adequação",
            descricao="Há lacunas relevantes. Priorize as ações pendentes, formalize evidências e defina responsáveis e prazos.",
            pacote="Pacote de Adequação NR-1",
            cor=BRAND["terracotta"],
        )
    return Diagnosis(
        pontos=pontos,
        maximo=maximo,
        percentual=percentual,
        status="CRÍTICO",
        status_label="Crítico",
        descricao="Alto risco de autuação. Recomenda-se ação imediata, diagnóstico completo e plano de adequação acompanhado.",
        pacote="Diagnóstico Completo + Plano de Ação Prioritário",
        cor=BRAND["danger"],
    )


def action_hint(resposta: str) -> str:
    if resposta == "Não - Inconformidade":
        return "Implementar requisito, definir responsável, prazo e evidências de comprovação."
    if resposta == "Em andamento":
        return "Concluir implementação, documentar evidências e atualizar responsável/prazo."
    if resposta == "N/A":
        return "Validar justificativa de não aplicabilidade e manter registro da decisão."
    return ""


def make_certificate_pdf(nome: str, certificado_id: str, created_at_br: str) -> bytes:
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=landscape(A4))
    width, height = landscape(A4)

    graphite = colors.HexColor(BRAND["graphite"])
    sand = colors.HexColor(BRAND["sand"])
    mist = colors.HexColor(BRAND["mist"])
    green = colors.HexColor(BRAND["green"])

    c.setFillColor(mist)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    c.setFillColor(graphite)
    c.rect(0, height - 3.4 * cm, width, 3.4 * cm, fill=1, stroke=0)
    if LOGO_SYMBOL_WHITE.exists():
        c.drawImage(ImageReader(str(LOGO_SYMBOL_WHITE)), 1.45 * cm, height - 2.65 * cm, width=1.35 * cm, height=1.75 * cm, mask="auto")
    if LOGO_WORDMARK_WHITE.exists():
        c.drawImage(ImageReader(str(LOGO_WORDMARK_WHITE)), 3.05 * cm, height - 2.65 * cm, width=6.2 * cm, height=1.25 * cm, mask="auto")

    c.setFillColor(sand)
    c.setFont("Times-Bold", 15)
    c.drawRightString(width - 1.5 * cm, height - 1.65 * cm, "CERTIFICADO DE PARTICIPAÇÃO")
    c.setFillColor(mist)
    c.setFont("Helvetica", 8)
    c.drawRightString(width - 1.5 * cm, height - 2.15 * cm, "Workshop NR-1 | Gestão de Riscos Psicossociais")

    c.setFillColor(colors.Color(0.35, 0.36, 0.40, alpha=0.08))
    c.circle(width - 5.0 * cm, height - 9.8 * cm, 3.5 * cm, fill=1, stroke=0)
    c.circle(width - 2.25 * cm, height - 9.8 * cm, 3.5 * cm, fill=1, stroke=0)

    c.setStrokeColor(graphite)
    c.setLineWidth(1.4)
    c.rect(1.1 * cm, 1.1 * cm, width - 2.2 * cm, height - 4.85 * cm, fill=0, stroke=1)
    c.setStrokeColor(sand)
    c.setLineWidth(0.9)
    c.rect(1.35 * cm, 1.35 * cm, width - 2.7 * cm, height - 5.35 * cm, fill=0, stroke=1)

    c.setFillColor(green)
    c.setFont("Helvetica", 13)
    c.drawCentredString(width / 2, height - 6.7 * cm, "CERTIFICAMOS QUE")

    c.setFont("Times-Bold", 34)
    c.drawCentredString(width / 2, height - 8.05 * cm, nome)

    c.setStrokeColor(sand)
    c.setLineWidth(1.0)
    c.line(width / 2 - 5.9 * cm, height - 8.45 * cm, width / 2 + 5.9 * cm, height - 8.45 * cm)

    c.setFillColor(graphite)
    c.setFont("Helvetica", 11)
    c.drawCentredString(width / 2, height - 9.15 * cm, "participou e concluiu com êxito o")

    c.setFont("Helvetica-BoldOblique", 15)
    c.drawCentredString(width / 2, height - 9.95 * cm, WORKSHOP["titulo"])

    c.setFont("Helvetica", 10)
    text = c.beginText(width / 2 - 9.2 * cm, height - 10.95 * cm)
    text.setLeading(15)
    text.textLine("promovido por PS&S Advogados, com conteúdo abrangendo as obrigações legais da NR-1")
    text.textLine("relativas à gestão de riscos psicossociais, ao GRO/PGR e às estratégias de prevenção")
    text.textLine("e bem-estar no trabalho.")
    c.drawText(text)

    y = height - 13.25 * cm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(sand)
    c.drawCentredString(width / 2 - 4.5 * cm, y, "CARGA HORÁRIA")
    c.drawCentredString(width / 2, y, "DATA")
    c.drawCentredString(width / 2 + 4.5 * cm, y, "LOCAL")
    c.setFillColor(graphite)
    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(width / 2 - 4.5 * cm, y - 0.45 * cm, WORKSHOP["carga_horaria"])
    c.drawCentredString(width / 2, y - 0.45 * cm, WORKSHOP["data"])
    c.drawCentredString(width / 2 + 4.5 * cm, y - 0.45 * cm, WORKSHOP["local"])

    sig_y = 2.7 * cm
    c.setStrokeColor(graphite)
    c.setLineWidth(0.6)
    c.line(4.3 * cm, sig_y + 0.8 * cm, 9.2 * cm, sig_y + 0.8 * cm)
    c.line(width - 9.2 * cm, sig_y + 0.8 * cm, width - 4.3 * cm, sig_y + 0.8 * cm)
    c.setFillColor(graphite)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(6.75 * cm, sig_y + 0.3 * cm, WORKSHOP["assinatura_1_nome"])
    c.drawCentredString(width - 6.75 * cm, sig_y + 0.3 * cm, WORKSHOP["assinatura_2_nome"])
    c.setFont("Helvetica", 7)
    c.drawCentredString(6.75 * cm, sig_y, WORKSHOP["assinatura_1_cargo"])
    c.drawCentredString(width - 6.75 * cm, sig_y, WORKSHOP["assinatura_2_cargo"])

    c.setFont("Helvetica", 7)
    c.setFillColor(graphite)
    c.drawString(1.7 * cm, 0.75 * cm, f"PS&S Advogados — {WORKSHOP['titulo']} | Emitido em {created_at_br} | ID: {certificado_id}")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer.getvalue()


def make_excel(participant: dict[str, str], answers: list[dict[str, Any]], summary: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_mat = wb.create_sheet("Matriz de Ações")
    ws_acao = wb.create_sheet("Plano de Ação")

    graphite = "595C65"
    sand = "D2C3A0"
    mist = "EDEDEC"
    green = "2B3928"
    white = "FFFFFF"
    thin = Side(style="thin", color="D7D7D7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws):
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

    ws_resumo.merge_cells("A1:B1")
    ws_resumo["A1"] = "CHECKLIST NR-1 — RESUMO DO DIAGNÓSTICO"
    ws_resumo["A1"].fill = PatternFill("solid", fgColor=graphite)
    ws_resumo["A1"].font = Font(color=white, bold=True, size=14)
    rows = [
        ("ID da submissão", summary["submission_id"]),
        ("ID do certificado", summary["certificado_id"]),
        ("Momento de preenchimento", summary["created_at_br"]),
        ("Nome", participant.get("nome", "")),
        ("E-mail", participant.get("email", "")),
        ("Empresa", participant.get("empresa", "")),
        ("CNPJ", participant.get("cnpj", "")),
        ("Cargo", participant.get("cargo", "")),
        ("Telefone", participant.get("telefone", "")),
        ("Pontuação Obtida", summary["pontuacao"]),
        ("Pontuação máxima", summary["pontuacao_maxima"]),
        ("Percentual de Conformidade", f"{summary['percentual']:.0f}%"),
        ("Status", summary["status"]),
        ("Pacote Recomendado", summary["pacote_recomendado"]),
    ]
    for idx, (k, v) in enumerate(rows, start=3):
        ws_resumo.cell(idx, 1, k)
        ws_resumo.cell(idx, 2, v)
        ws_resumo.cell(idx, 1).font = Font(bold=True, color=green)
        if k in ["Pontuação Obtida", "Percentual de Conformidade", "Status", "Pacote Recomendado"]:
            ws_resumo.cell(idx, 1).fill = PatternFill("solid", fgColor=mist)
            ws_resumo.cell(idx, 2).fill = PatternFill("solid", fgColor=sand)
            ws_resumo.cell(idx, 2).font = Font(bold=True, color="000000")
    ws_resumo.column_dimensions["A"].width = 34
    ws_resumo.column_dimensions["B"].width = 68

    headers = [
        "Área", "Nº", "Requisito", "Pergunta de Verificação", "Resposta", "Pontuação",
        "Observação", "Ação Necessária", "Responsável pela Ação", "Prazo", "Status da Ação", "Observações Internas",
    ]
    ws_mat.append(headers)
    for a in answers:
        ws_mat.append([
            a["area"], a["numero"], a["requisito"], a["pergunta"], a["resposta"], a["pontuacao"],
            a.get("observacao", ""), a.get("acao_sugerida", ""), "", "", "", "",
        ])

    headers_acao = ["Nº do Item", "Requisito", "Pergunta", "Resposta", "Ação Necessária", "Responsável", "Prazo", "Status", "Observações"]
    ws_acao.append(headers_acao)
    for a in answers:
        if a["resposta"] in ["Não - Inconformidade", "Em andamento", "N/A"]:
            ws_acao.append([a["numero"], a["requisito"], a["pergunta"], a["resposta"], a["acao_sugerida"], "", "", "", a.get("observacao", "")])

    for ws in [ws_mat, ws_acao]:
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=graphite)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    for row in ws_mat.iter_rows(min_row=2):
        resp = row[4].value
        fill = None
        if resp == "Sim - Conforme":
            fill = PatternFill("solid", fgColor="DDEBDD")
        elif resp == "Em andamento":
            fill = PatternFill("solid", fgColor="FFF2CC")
        elif resp == "Não - Inconformidade":
            fill = PatternFill("solid", fgColor="F4CCCC")
        elif resp == "N/A":
            fill = PatternFill("solid", fgColor="E8E6DE")
        if fill:
            for cell in row:
                cell.fill = fill

    for ws in [ws_resumo, ws_mat, ws_acao]:
        style_sheet(ws)
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 46)
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 24

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def render_header() -> None:
    symbol = asset_b64(LOGO_SYMBOL_WHITE)
    wordmark = asset_b64(LOGO_WORDMARK_WHITE)

    symbol_html = (
        f'<img class="pss-logo-symbol" src="data:image/png;base64,{symbol}" alt="Logo PS&S">'
        if symbol
        else ""
    )

    wordmark_html = (
        f'<img class="pss-logo-wordmark" src="data:image/png;base64,{wordmark}" alt="Perroni Sanvicente & Schirmer">'
        if wordmark
        else '<div class="pss-wordmark-fallback">PERRONI SANVICENTE<br>&amp; SCHIRMER</div>'
    )

    st.markdown(
        f"""<header class="pss-hero">
<div class="pss-hero-inner">
<div class="pss-brand-lockup">
{symbol_html}
{wordmark_html}
</div>
<div class="pss-hero-copy">
<div class="pss-hero-title">Checklist de Adequação à NR-1</div>
<p>Diagnóstico digital para mapeamento de conformidade em riscos psicossociais, GRO/PGR, prevenção, documentação e política interna.</p>
</div>
</div>
</header>""",
        unsafe_allow_html=True,
    )


def render_intro() -> None:
    st.markdown(
        """
        <div class="pss-card pss-intro">
            <p><strong>Como preencher:</strong> para cada requisito, selecione uma das quatro opções — <strong>Sim - Conforme</strong> (5 pontos), <strong>Em andamento</strong> (2,5 pontos), <strong>Não - Inconformidade</strong> (0 pontos) ou <strong>N/A</strong> (não se aplica). Ao final, clique em <strong>Calcular e Enviar Resultados</strong>.</p>
            <p style="margin-bottom:0"><strong>Responda, gere seu score, Matriz de Ações e emita o seu certificado de participação no Workshop!</strong></p>
            <p class="small-note" style="margin-bottom:0;margin-top:10px">A observação é opcional, mas recomendamos justificar quando a resposta for N/A.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def init_state() -> None:
    defaults = {
        "identificado": False,
        "participant": {},
        "start_id": "",
        "result_payload": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def render_identification_gate() -> bool:
    if st.session_state.identificado:
        return True

    st.markdown('<div class="pss-card"><div class="pss-card-title">Identificação da Empresa</div>', unsafe_allow_html=True)
    with st.form("identificacao_form", clear_on_submit=False):
        c1, c2 = st.columns(2)
        with c1:
            nome = st.text_input("Nome completo *", value=st.session_state.participant.get("nome", ""))
            email = st.text_input("E-mail *", value=st.session_state.participant.get("email", ""))
            cargo = st.text_input("Cargo", value=st.session_state.participant.get("cargo", ""))
        with c2:
            empresa = st.text_input("Empresa *", value=st.session_state.participant.get("empresa", ""))
            cnpj = st.text_input("CNPJ", value=st.session_state.participant.get("cnpj", ""))
            telefone = st.text_input("Telefone", value=st.session_state.participant.get("telefone", ""))

        aceite = st.checkbox(
            "Declaro estar ciente de que meus dados e respostas serão usados para registro de participação, emissão do certificado e geração do diagnóstico NR-1.",
            value=False,
        )
        start = st.form_submit_button("Salvar dados e iniciar questionário")
    st.markdown("</div>", unsafe_allow_html=True)

    if start:
        participant = {
            "nome": nome.strip(),
            "email": email.strip(),
            "empresa": empresa.strip(),
            "cnpj": cnpj.strip(),
            "cargo": cargo.strip(),
            "telefone": telefone.strip(),
        }
        missing = [label for label, value in {"Nome completo": nome, "E-mail": email, "Empresa": empresa}.items() if not value.strip()]
        if missing:
            st.error("Preencha os campos obrigatórios: " + ", ".join(missing))
            return False
        if not aceite:
            st.error("É necessário marcar o aceite de processamento de dados para liberar o questionário.")
            return False
        start_id = "START-" + uuid.uuid4().hex[:10].upper()
        now_utc = datetime.now(timezone.utc)
        now_br = datetime.now(TZ_BR)
        save_start_record(participant, start_id, now_utc, now_br.strftime("%d/%m/%Y %H:%M:%S"))
        st.session_state.identificado = True
        st.session_state.participant = participant
        st.session_state.start_id = start_id
        st.rerun()

    st.info("O questionário será exibido após o preenchimento da identificação e aceite de processamento de dados.")
    return False


def render_company_summary(participant: dict[str, str]) -> None:
    def item(label: str, value: str) -> str:
        return f'<div class="pss-company-item"><span class="label">{escape(label)}</span><span class="value">{escape(value or "-")}</span></div>'

    st.markdown(
        f"""
        <div class="pss-card">
            <div class="pss-card-title">Identificação da Empresa</div>
            <div class="pss-company-card">
                {item("Nome", participant.get("nome", ""))}
                {item("E-mail", participant.get("email", ""))}
                {item("Empresa", participant.get("empresa", ""))}
                {item("CNPJ", participant.get("cnpj", ""))}
                {item("Cargo", participant.get("cargo", ""))}
                {item("Telefone", participant.get("telefone", ""))}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def progress_html(respondidas: int, total: int) -> None:
    st.markdown(
        f"""
        <div class="pss-progress">
            <span>Perguntas respondidas:</span>
            <span>{respondidas}/{total}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.progress(respondidas / total if total else 0)


def section_ranges(questions: list[Question]) -> dict[str, tuple[int, int]]:
    ranges: dict[str, list[int]] = {}
    for q in questions:
        ranges.setdefault(q.area, []).append(q.numero)
    return {area: (min(nums), max(nums)) for area, nums in ranges.items()}


def render_questions(questions: list[Question]) -> tuple[dict[int, str | None], dict[int, str]]:
    responses: dict[int, str | None] = {}
    notes: dict[int, str] = {}
    ranges = section_ranges(questions)
    current_area = None

    for q in questions:
        if q.area != current_area:
            current_area = q.area
            start, end = ranges[q.area]
            st.markdown(
                f"""
                <div class="pss-section-header">
                    <span class="name">{escape(q.area)}</span>
                    <span class="count">Itens {start} a {end}</span>
                </div>
                """,
                unsafe_allow_html=True,
            )
        with st.container(border=True):
            st.markdown(
                f"""
                <div class="pss-question-heading">
                    <div class="pss-question-meta">
                        <span class="pss-question-num">{q.numero}</span>
                        <span class="pss-question-req">{escape(q.requisito)}</span>
                    </div>
                    <p class="pss-question-text">{escape(q.pergunta)}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            responses[q.numero] = st.radio(
                label=f"Resposta do item {q.numero}",
                options=STATUS_OPTIONS,
                index=None,
                horizontal=True,
                key=f"q_{q.numero}",
            )
            notes[q.numero] = st.text_area(
                "Observação",
                key=f"obs_{q.numero}",
                placeholder="Opcional. Recomendamos justificar quando a resposta for N/A.",
                height=70,
            )
    return responses, notes


def build_answers(questions: list[Question], participant: dict[str, str], responses: dict[int, str | None], notes: dict[int, str], summary_ids: dict[str, str]) -> list[dict[str, Any]]:
    answers: list[dict[str, Any]] = []
    for q in questions:
        resposta = responses.get(q.numero)
        if resposta is None:
            continue
        score = SCORES[resposta]
        observacao = notes.get(q.numero, "").strip()
        answers.append({
            "submission_id": summary_ids["submission_id"],
            "certificado_id": summary_ids["certificado_id"],
            "created_at": summary_ids["created_at"],
            "nome": participant.get("nome", ""),
            "email": participant.get("email", ""),
            "empresa": participant.get("empresa", ""),
            "numero": q.numero,
            "area": q.area,
            "requisito": q.requisito,
            "pergunta": q.pergunta,
            "resposta": resposta,
            "pontuacao": score,
            "evidencia_observacao": observacao,
            "observacao": observacao,
            "acao_sugerida": action_hint(resposta),
        })
    return answers


def render_result(payload: dict[str, Any]) -> None:
    diagnosis: Diagnosis = payload["diagnosis"]
    pct = int(diagnosis.percentual)
    st.markdown(
        f"""
        <div class="pss-result" style="--status-color:{diagnosis.cor};">
            <div class="pss-result-kicker">Resultado do Diagnóstico</div>
            <div class="pss-result-points">{diagnosis.pontos:g}</div>
            <div class="pss-result-max">/ {diagnosis.maximo:g} pontos</div>
            <div class="pss-result-bar">
                <div class="pss-result-bar-fill" style="--value:{pct}%;"></div>
            </div>
            <div class="pss-status-pill">{escape(diagnosis.status_label)}</div>
            <div class="pss-result-copy">
                <p><strong>Percentual de Conformidade:</strong> {pct}%</p>
                <p><strong>Pacote Recomendado:</strong> {escape(diagnosis.pacote)}</p>
                <p>{escape(diagnosis.descricao)}</p>
            </div>
            <div class="pss-cert-id">ID do certificado: {escape(payload["certificado_id"])}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not payload.get("save_ok", True):
        st.warning("Não foi possível confirmar o registro remoto. Os arquivos foram gerados normalmente.")


    c0, c1, c2, c3 = st.columns([1.0, 1.25, 1.8, 1.0])
    with c1:
        st.download_button(
            "Download do certificado",
            data=payload["pdf_bytes"],
            file_name=f"certificado_{payload['certificado_id']}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    with c2:
        st.download_button(
            "Download dos resultados e Matriz de Ações",
            data=payload["excel_bytes"],
            file_name=payload["excel_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def footer() -> None:
    st.markdown(
        """
        <footer class="pss-footer-bleed">
            <div class="pss-footer-inner">
                <div class="brand">PS&amp;S Advogados</div>
                <div class="info">carolina@pssadv.com.br &nbsp;|&nbsp; www.pssadv.com.br</div>
                <div class="muted">Documento de uso confidencial — NR-1 Compliance · Portaria MTE n. 1.419/2024</div>
            </div>
        </footer>
        """,
        unsafe_allow_html=True,
    )


def reset_state() -> None:
    for key in list(st.session_state.keys()):
        if key.startswith("q_") or key.startswith("obs_"):
            del st.session_state[key]
    st.session_state.identificado = False
    st.session_state.participant = {}
    st.session_state.start_id = ""
    st.session_state.result_payload = None
    st.rerun()


def main() -> None:
    st.set_page_config(
        page_title="Checklist NR-1 | PS&S Advogados",
        page_icon="○",
        layout="wide",
    )
    init_state()
    css()
    render_header()

    st.markdown('<main class="pss-content">', unsafe_allow_html=True)

    if st.session_state.result_payload:
        render_company_summary(st.session_state.participant)
        render_result(st.session_state.result_payload)
        n0, n1, n2 = st.columns([1.6, 1.0, 1.6])
        with n1:
            if st.button("Novo preenchimento", use_container_width=True):
                reset_state()
        st.markdown("</main>", unsafe_allow_html=True)
        footer()
        return

    render_intro()

    questions = load_questions()
    if not render_identification_gate():
        st.markdown("</main>", unsafe_allow_html=True)
        footer()
        return

    participant = st.session_state.participant
    responses, notes = render_questions(questions)
    respondidas = sum(1 for v in responses.values() if v is not None)
    progress_html(respondidas, len(questions))

    c1, c2 = st.columns([3, 1])
    with c1:
        submit = st.button("Calcular e Enviar Resultados", type="primary", use_container_width=True)
    with c2:
        reset = st.button("Recomeçar", use_container_width=True)

    if reset:
        reset_state()

    if submit:
        missing_questions = [q.numero for q in questions if responses.get(q.numero) is None]
        if missing_questions:
            st.error("Responda todas as perguntas. Pendentes: " + ", ".join(map(str, missing_questions)))
        else:
            submission_id = str(uuid.uuid4())
            certificado_id = "PSS-NR1-" + uuid.uuid4().hex[:8].upper()
            now_utc = datetime.now(timezone.utc)
            now_br = datetime.now(TZ_BR)
            created_at_br = now_br.strftime("%d/%m/%Y %H:%M:%S")
            summary_ids = {
                "submission_id": submission_id,
                "certificado_id": certificado_id,
                "created_at": now_utc.isoformat(),
            }
            answers = build_answers(questions, participant, responses, notes, summary_ids)
            diagnosis = calculate(answers)
            summary = {
                "submission_id": submission_id,
                "start_id": st.session_state.start_id,
                "certificado_id": certificado_id,
                "created_at": now_utc.isoformat(),
                "created_at_br": created_at_br,
                "pontuacao": diagnosis.pontos,
                "pontuacao_maxima": diagnosis.maximo,
                "percentual": diagnosis.percentual,
                "status": diagnosis.status,
                "pacote_recomendado": diagnosis.pacote,
            }
            submission = {
                **summary,
                **participant,
                "workshop_titulo": WORKSHOP["titulo"],
                "respostas_json": json.dumps(answers, ensure_ascii=False, default=str),
                "user_agent": "streamlit",
                "fonte": "app_streamlit",
            }
            save_ok, _ = save_to_bigquery(submission, answers)

            pdf_bytes = make_certificate_pdf(participant["nome"], certificado_id, created_at_br)
            excel_bytes = make_excel(participant, answers, summary)
            safe_empresa = participant["empresa"].strip().replace(" ", "_").replace("/", "-")
            st.session_state.result_payload = {
                "diagnosis": diagnosis,
                "certificado_id": certificado_id,
                "pdf_bytes": pdf_bytes,
                "excel_bytes": excel_bytes,
                "excel_name": f"matriz_acoes_nr1_{safe_empresa}_{certificado_id}.xlsx",
                "save_ok": save_ok,
            }
            st.rerun()

    st.markdown("</main>", unsafe_allow_html=True)
    footer()


if __name__ == "__main__":
    main()
